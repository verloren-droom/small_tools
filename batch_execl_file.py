from argparse import ArgumentParser
from json import load, dump
from shutil import copy
from pyperclip import paste
from openpyxl import load_workbook
from tkinter import messagebox, Tk

import sys
from os import path
from platform import system

is_macos = system() == 'Darwin'
is_release = getattr(sys, 'frozen', False)

def is_within_app_bundle(executable_path):
    if not is_macos:
        return False
    return ".app/" in executable_path

def get_default_file_name():
    if is_release:
        if is_macos:
            executable_path = path.join(sys.executable, '../../MacOS/' + path.basename(sys.executable))
            return path.splitext(path.basename(executable_path))[0]
        else:
            return path.splitext(path.basename(sys.executable))[0]
    else:
        return path.splitext(path.basename(path.abspath(__file__)))[0]

def get_default_abs_path():
    if is_release:
        if is_macos:
            is_in_app_bundle = is_within_app_bundle(sys.executable)
            if is_in_app_bundle:
                app_bundle_path = path.dirname(path.dirname(path.dirname(sys.executable)))
                return path.dirname(app_bundle_path) + "/"
            else:
                return path.dirname(sys.executable) + "/"
        else:
            return path.dirname(sys.executable) + "/"
    else:
        return path.dirname(path.abspath(__file__)) + "/"


DEFAULT_CONFIG = {
    "path": f"{get_default_abs_path()}{get_default_file_name()}",
    "content": paste().strip(),
    "start": "2:5",
    "end": ":5",
    "copy": False,
    "tip" : True
}

def create_config_template(output_path):
    with open(output_path, "w", encoding='utf-8') as f:
        dump(DEFAULT_CONFIG, f, indent=4, ensure_ascii=False)

def read_config_file(config_file_path):
    try:
        with open(config_file_path, "r", encoding='utf-8') as f:
            config = load(f)
            return config
    except FileNotFoundError:
        return DEFAULT_CONFIG
    except Exception:
        quit_batch()

def get_index_rc(ws, pos, showtip=False):
    if pos == "":
        quit_batch()

    if pos.startswith(":"):
        if pos.endswith(":"):
            return ws.max_row, ws.max_column
        else:
            col = int(pos[1:])
            return ws.max_row, col if col > 0 else ws.max_column

    if pos.endswith(":"):
        row = int(pos[:-1])
        return row, ws.max_column if row > 0 else 1

    if ":" in pos:
        row, col = map(int, pos.split(":"))
        return row, col if row > 0 and col > 0 else 1

    try:
        num = int(pos)
        return num, num if num > 0 else 1
    except ValueError:
        if (showtip):
            messagebox.showwarning("Input Error", f"Input {pos} format error, please use Row:Column\nRow:[1,] and Column:[1,]!")
        quit_batch()

def copy_file(excel_file_path):
    base_name, ext = path.splitext(excel_file_path)
    new_file_path = base_name + "_Copy" + ext
    copy(excel_file_path, new_file_path)
    return new_file_path

def quit_batch():
    sys.exit()

def main(args):
    if args.cc and not path.exists(args.config):
        create_config_template(args.config)
        quit_batch()
    config = read_config_file(args.config)

    tip = args.t if args.t else config['tip'] 

    if tip:
        root = Tk()
        root.withdraw()
    
    excel_file = args.p if args.p else config["path"]

    if (args.copy if args.copy else config["copy"]):
        excel_file = copy_file(excel_file)

    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except FileNotFoundError:
        if tip:
            messagebox.showwarning("File not found", f"{excel_file} is not exist!")
        quit_batch()
    except Exception:
        quit_batch()

    content = args.c if args.c else config["content"]
    if not content:
        if tip:
            messagebox.showwarning("Content Error", "clipboard or -c parameter is null")
        quit_batch()
    
    start_cell = args.s if args.s else config["start"]
    end_cell = args.e if args.e else config["end"]
    start_row, start_col = get_index_rc(ws, start_cell, tip)
    end_row, end_col = get_index_rc(ws, end_cell, tip)

    if tip and not messagebox.askyesno("Write content", f"modify {min(start_row, end_row)}:{max(start_row, end_row) + 1} to {min(start_col, end_col)}:{max(start_col, end_col) + 1}\ncontent: {content if len(content) < 100 else content[:100] + ' ......'}\nWhether to write?"):
        messagebox.showwarning("Write interrupts", f"stop writing to {excel_file}")
        quit_batch()

    for row in range(min(start_row, end_row), max(start_row, end_row) + 1):
        for col in range(min(start_col, end_col), max(start_col, end_col) + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and cell.value != "":
                cell.value = content

    wb.save(excel_file)
    if tip:
        messagebox.showinfo("Write successfull", f"write to {excel_file}")
        root.destroy()

if __name__ == "__main__":
    parser = ArgumentParser(description="""
                            Batch modification of Excel (*.xlsx) files,
                            URL: [https://github.com/verloren-droom/small_tools/blob/main/batch_execl_file.py]
                            """)

    parser.add_argument("-p", help=f"excel (.xlsx) file path, default: {get_default_file_name()}.xlsx", default=f"{DEFAULT_CONFIG['path']}.xlsx")
    parser.add_argument("-c", help=f"content to modify, default: {DEFAULT_CONFIG['content'] if len(DEFAULT_CONFIG['content']) < 50 else DEFAULT_CONFIG['content'][:50] + '  ......'}", default=DEFAULT_CONFIG["content"])
    parser.add_argument("-s", help=f"starting cell, format Row:Column. Default: {DEFAULT_CONFIG['start']}", default=DEFAULT_CONFIG["start"])
    parser.add_argument("-e", help=f"ending cell, format Row:Column. Default: {DEFAULT_CONFIG['end']}", default=DEFAULT_CONFIG["end"])
    parser.add_argument("-t", action="store_true", help="show pop-up prompt")
    parser.add_argument("-copy", action="store_true", help=f"create a copy of the Excel ({get_default_file_name()}_Copy.xlsx) file")
    parser.add_argument("-cc", action="store_true", help=f"create a JSON configuration file template (skipped if already exists), default: {get_default_file_name()}.json")
    parser.add_argument("-config", help=f"JSON configuration file, default: {get_default_file_name()}.json", default=f"{get_default_abs_path()}{get_default_file_name()}.json")

    args = parser.parse_args()

    main(args)